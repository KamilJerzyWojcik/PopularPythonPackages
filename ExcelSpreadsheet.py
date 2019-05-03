import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)


sheet = wb["Sheet1"]
cell = sheet["a1"]
cell2 = sheet.cell(row=1, column=1)

sheet.append([1, 2, 3])
wb.save("transactions2.xlsx")


# column = sheet["a"]
# print(column)
# print("")
# column = sheet["a:c"]
# print(column)
# column = sheet["a1:c3"]
# print(column)

# range_cells = sheet[1:3]
# print(range_cells)

# print(sheet.max_row)
# print(sheet.max_column)

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# cell.value = 123
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)



# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)